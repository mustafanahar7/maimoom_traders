from django.shortcuts import render , HttpResponse,redirect 
from django.http import JsonResponse
from .models import CustomUser , ProductInventory , salesCustomerData,customers_order_from_website,website_orderItems , Queryform

# Create your views here.
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.db import IntegrityError
from django.contrib.auth import authenticate, login , logout
from django.contrib.auth.decorators import login_required
import os
from django.conf import settings
from django.template.loader import render_to_string
import pdfkit
import tempfile
import io
from openpyxl import Workbook
from django.db.models import Sum
from django.core.mail import send_mail,EmailMessage
from django.utils.html import strip_tags
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.conf import settings
import math
import razorpay
from datetime import date , timedelta
from django.db.models import Count



# from django_weasyprint import WeasyTemplateResponse

        
def splash_page(request):
    return render(request,'splashpage.html')  
   


@csrf_exempt
def registerPage(request):
    message=None
    error_message=None
    if request.method=='POST':
        first_name= request.POST.get('first_name')
        last_name= request.POST.get('last_name')
        phone_number= request.POST.get('phone')
        profile= request.FILES.get('profile')
        email= request.POST.get('email')
        password_val = request.POST.get('password')
        
        # print(f'name -- {first_name +last_name} , phone -- {phone_number} , email -- {email}')
        
        try:
            store_user = CustomUser.objects.create_user(
                first_name=first_name,
                last_name=last_name,
                email=email,
                mobile_number=phone_number,
                password=password_val,
                profile_img = profile
            )
            message = 'Account created successfully!'
            return redirect('loginpage')     
        except IntegrityError:
            error_message = 'Email address already exists. Please use a different email.'
    return render(request, 'registerpage.html',{'error_message':error_message,'messages':message})



@csrf_exempt
def siginpage(request):
    message=None
    error_message=None
    if request.method=='POST':
        username = request.POST.get('email')
        password_val=request.POST.get('password')
        
        user = authenticate(email=username,password=password_val)
        
        print(user)
        if user is not None and user.is_superuser==True:
            login(request,user)        
            request.session['username'] = username
            message = 'login successfully!'
            return redirect('admin_dashboard')
        
        elif user is not None:
            login(request,user)
            request.session['username'] = username
            message =  'login successfully!'
            return redirect('homepage')
        
        else:
            error_message = 'check your credential.'
            
    return render(request, 'loginpage.html',{'error_message':error_message,'messages':message})
        
   
  
        
        
@login_required      
def homePage(request):
   
    message_list = messages.get_messages(request)
    username = request.user
    print('<<<<<<<<<<<<<<<<<<<')
    print(username)
    print('<<<<<<<<<<<<<<<<<<<')
    get_user_details = CustomUser.objects.get(email=username)
    print(get_user_details.profile_img)
    oos=None
    
    ######## get the all pic for slider for home page ######
    directory = 'D:/personal proj/project1/maimoontraders/media/slider_images'
    file_names = os.listdir(directory)
    files=[]

    for file_name in file_names:
        files.append(file_name)
        
    # messages for slider
    slider_msg = ['Welcome to Maimoon Traders','Explore your Desire','Thank For Visiting','New Collection']
    
    files = zip(files,slider_msg)
    ##################################################################
    
    search_val = request.GET.get('q')
    if request.method=="GET":
        if search_val:
            fetch_data = ProductInventory.objects.filter(product_name__icontains=search_val)
        else:  
            fetch_data = ProductInventory.objects.filter(qty_sqft__gt=0)
    return render(request , 'homepage.html',{'imagedata':fetch_data,'images':files,'msg':slider_msg,'oos':oos,'message_list': message_list,'user_profile':get_user_details})





@login_required
def signout(request):
    logout(request)
    return redirect('loginpage')





@csrf_exempt
@login_required 
def uploadphoto(request):
    current_path = request.path
    error_message=None
    message=None
    
    if request.method=="POST":
        
        code = request.POST.get('code')
        name = request.POST.get('name')
        type = request.POST.get('type')
        category = request.POST.get('category')
        description = request.POST.get('description')
        qty = request.POST.get('quantity')
        qty_ft = request.POST.get('qty_sqft')
        image_file = request.FILES.get('image')
        sale_price = request.POST.get('sale_price')
        cost = request.POST.get('cost')
        godown = request.POST.get('godown')
        total_amount = int(qty_ft)* float(cost)*int(qty)
        
        # print('cost is -----',cost)
        current_date = date.today()
        
        try:
            get_product_obj = ProductInventory.objects.get(product_code=code)
            
            print('=================================')
            print(get_product_obj.product_image)
            print('=================================')
            
            get_product_obj.product_code = code
            get_product_obj.product_name = name
            get_product_obj.product_type=type
            get_product_obj.product_category = category
            # get_product_obj.product_description = description,
            get_product_obj.quantity += int(qty)
            get_product_obj.product_image = image_file
            get_product_obj.cost_sqft = cost
            get_product_obj.selling_price_sqft = sale_price
            get_product_obj.total_amount_sqft  = float(get_product_obj.cost_sqft) * float(get_product_obj.qty_sqft+float(qty_ft)) 
            get_product_obj.godown_name = godown 
            get_product_obj.qty_sqft += float(qty_ft)
            get_product_obj.last_purchase_date=current_date
            
            get_product_obj.save()
            message =  'Product update successfully!'
            
            
        except ProductInventory.DoesNotExist:
            store_product = ProductInventory.objects.create(
                                            product_code = code,
                                            product_name = name,
                                            product_type=type,
                                            product_category = category,
                                            product_description = description,
                                            quantity = qty,
                                            product_image = image_file,
                                            cost_sqft = cost,
                                            selling_price_sqft = sale_price,
                                            total_amount_sqft = total_amount,
                                            godown_name = godown ,
                                            qty_sqft = qty_ft,
                                            first_purchase_date = current_date,
                                            last_purchase_date=current_date                                      
            )
            message =  'Product Stored successfully!'
            
        

    return render(request, 'uploadimage.html',{'current_path':current_path,'error_message':error_message,'message':message})




 
def adminpage(request):
    
    adminuser = request.user
    ####### check if user is superuser then only allow to admin page 
    try :
        admin_obj = CustomUser.objects.get(email=adminuser)
        if admin_obj.is_superuser==True:
            return render (request, 'adminDashboard.html')
        else:
            return redirect('loginpage')
    except Exception as e:
        return redirect('registerpage')
        
 


@login_required 
def categoryPage(request):
    distinct_values = ProductInventory.objects.values_list('product_category', flat=True).distinct()
    username = request.user
    print('<<<<<<<<<<<<<<<<<<<')
    print(username)
    print('<<<<<<<<<<<<<<<<<<<')
    get_user_details = CustomUser.objects.get(email=username)
    print(get_user_details.profile_img)
    print('>>>>>>>',distinct_values)

    fetch_data = ProductInventory.objects.all()
    return render(request,'category.html',{'list_of_cat':distinct_values,'products':fetch_data,'user_profile':get_user_details})


@login_required 
def categoryfilter(request,product_category):
    # active_category = "All Products"
    distinct_values = ProductInventory.objects.values_list('product_category', flat=True).distinct()
    active_category = distinct_values
    print('>>>>>>>',distinct_values)
    fetch_data = ProductInventory.objects.filter(product_category=product_category)

    return render(request,'category.html',{'list_of_cat':distinct_values,'products':fetch_data, 'active_category': active_category})





@login_required
def AdminDashboard(request):
    return render(request, 'adminDashboard.html')


@login_required 
def AdminProductList(request):
    all_product = ProductInventory.objects.all()
    
    if request.GET.get('download')=='xlsx':
        # Create a new workbook
        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        ws.append(['Product Code', 'Product Name', 'Description', 'Price' , 'Cost' ,'Total sqft','Godown'])

        # Add data rows to the worksheet
        for product in all_product:
            ws.append([product.product_code, product.product_name, product.product_description, product.selling_price_sqft , product.cost_sqft,product.qty_sqft,product.godown_name])

        # Save the workbook to a BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create an HTTP response with the Excel file as attachment
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="product_list.xlsx"'
        return response
   
    return render(request , 'adminproductlistpage.html',{'productList':all_product})




# global sale_list = []
# global sales_data=[]
@login_required 
def AdminSalesaPage(request):
    
    # if 'sale_list' not in request.session:
    #     request.session['sale_list'] = []
    
    # if 'sales_data' not in request.session:
    #     request.session['sales_data'] = []

    # sale_list = request.session['sale_list']
    # if 'sale_list' not in request.session:
    sales_data = request.session.get('sales_data', [])
    sale_list = request.session.get('sale_list', [])
    message=None
    error_message=None   
    current_date=date.today() 
    
    if request.method=="GET":
        product_code = request.GET.get('code')  # Get the product code from the GET request
        if product_code:
            try:
                products = ProductInventory.objects.get(product_code=product_code) 
                qty_in_godown = products.qty_sqft 
                print('<<<<<<<>>>>>>>>>>',products)
                
                return render(request, 'adminsalesdashboard.html', {'product': products,'qty_in_godown':qty_in_godown})
            except ProductInventory.DoesNotExist:
                error_message = 'Product not found.'
                return render(request, 'adminsalesdashboard.html', {'error_message': error_message})
        else:
            return render(request, 'adminsalesdashboard.html',{'add_items': sale_list})
        
        
    if request.method=="POST":
        print(';;;;;;;;;;;;;;;;;;;;;;;;;;')
        print(request.POST)
        button_pressed = request.POST.get('button_pressed')
        
        if request.POST.get('action') == 'delete':
            product_code_to_delete = request.POST.get('product_code')
            # Remove the product from sale_list
            sale_list = [item for item in request.session['sale_list'] if item['product_code'] != product_code_to_delete]
            request.session['sale_list'] = sale_list
            return render(request, 'adminsalesdashboard.html', {'add_items': sale_list})
        
        
        
        
        elif button_pressed=='add_button':
            print(';;;;;;;;;;;;;;;;;;;;;;;;;;')
            print("Add Button is Pressed")
            print(';;;;;;;;;;;;;;;;;;;;;;;;;;')
            product_obj = ProductInventory.objects.get(product_code=request.POST.get('code'))
            sale_list.append(
            {'product_code':request.POST.get('code'),
            'customer': request.POST.get('customer_name'),
            'customer_address': request.POST.get('customer_address'),
            'phone_number' : request.POST.get('phone_number'),
            'sqft_price' : request.POST.get('sale_price'),
            'qty' : float(request.POST.get('qty_sqft')),
            'email' : request.POST.get('email'),
            'discount' : request.POST.get('discount'),
            'final_price' : request.POST.get('final_price'),
            'total_amount' : request.POST.get('total_price'),
            'product_name':product_obj.product_name
            })
            
            request.session['sale_list'] = sale_list

            
            print(sale_list)
            return render(request,'adminsalesdashboard.html' ,{'add_items':sale_list})
        
        elif button_pressed=='selled':
            print('inside sale button press')
            sale_list = request.session['sale_list']
            print(sale_list)
            # product_code = request.POST.get('code')
            # customer = request.POST.get('customer_name')
            # customer_address = request.POST.get('customer_address')
            # phone_number = request.POST.get('phone_number')
            # sqft_price = request.POST.get('sale_price')
            # qty = float(request.POST.get('qty_sqft'))
            # email = request.POST.get('email')
            # discount = request.POST.get('discount')
            # final_price = request.POST.get('final_price')
            # total_amount = request.POST.get('total_price')
            total_bill_amount=0
            for item in sale_list:
                product_obj = ProductInventory.objects.get(product_code=item['product_code'])
                
                if product_obj and product_obj.qty_sqft>=item['qty']:
                    product_obj.qty_sqft =  product_obj.qty_sqft-float(item['qty'])
                    product_obj.total_amount_sqft -= (product_obj.cost_sqft)*float(item['qty']) 
                    product_obj.save()
                    
                                        
                    ######### store customer data #########
                    #  forproduct_code, qty ,sqft_price, total_amount in zip(product_code_list, qty_list,sqft_price_list, total_amount_list):
                    store_sales = salesCustomerData.objects.create(
                                                                        sales_product_id=item['product_code'],
                                                                        customer_name = item['customer'],
                                                                        customer_address=item['customer_address'],
                                                                        customer_phone_number = item['phone_number'],
                                                                        qty_in_sqft =item['qty'],
                                                                        sale_price_sqft = item['sqft_price'],
                                                                        total_sale_amount= item['final_price'],
                                                                        discount=item['discount']
                                                                        ) 
                    
                    total_bill_amount += float(item['final_price'])            
                    sales_data.append({
                            'customer_name':item['customer'],
                            'customer_address':item['customer_address'],
                            'customer_contact':item['phone_number'],
                            'product_name': product_obj.product_name,
                            'quantity': item['qty'],
                            'email':item['email'],
                            'unit_price': item['sqft_price'],
                            'total': item['total_amount'],
                            # 'total_amount': item['total_amount'], # Total amount for the entire bill
                            'date':current_date,
                            'discount':item['discount'],
                            'final_price':item['final_price']
                        })
            
                else:
                    error_message = 'Product may out of stock of less qty available'
            
            
                
                
                    

            message =  'Product saled successfully!'
            
            # Render HTML bill content using a template and sales data
            bill_content = render_to_string('bill_template.html', {'sales_data':sales_data,'total_amount':total_bill_amount})
            subject='Maimoon Traders Bill'
            print(f'mail sending')
            send_mail(
                    subject,
                    '',
                    from_email=settings.DEFAULT_FROM_EMAIL,  # Sender's email (set in your settings)
                    recipient_list=[sale_list[0]['email']],
                    html_message=bill_content,
                    fail_silently=False,  # Set to True to suppress exceptions if sending fails
                )
                # PDFKIT_CONFIG = pdfkit.configuration(wkhtmltopdf='C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf')
                # pdf = pdfkit.from_string(bill_content, False, configuration=PDFKIT_CONFIG)
            # request.session['sale_list'] = []
            # request.session['sales_data']=[]
            # sale_list=[]
            # sales_data=[]
            request.session['sale_list']=[]
            request.session['sales_data']=[]
            return HttpResponse(bill_content)
    
        
            
        
         
    return render(request, 'adminsalesdashboard.html',{'error_message':error_message,'message':message,'add_items': sale_list})
    


    
    
    
from datetime import date 
def Admin_Salesreport(request):
    
    current_date = str(date.today())
    # sales_report = salesCustomerData.objects.all()
    sales_report = customers_order_from_website.objects.all()
    
    if request.GET.get('download')=='xlsx':

        wb = Workbook()
        ws = wb.active

        # Add headers to the worksheet
        ws.append(['Sale Date','Order Number','Customer Name', 'Customer Address', 'Mobile Number'  ,'Total Bill Amount'])

        # Add data rows to the worksheet
        for sales in sales_report:
            ws.append([sales.created_at,sales.order_number,sales.customer_name, sales.address, sales.customer_phone, sales.total_amount])

        # Save the workbook to a BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create an HTTP response with the Excel file as attachment
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="sales_report {current_date}.xlsx"'
        print('sales file created ')
        return response
        
    return render (request , 'adminsalesreport.html',{'report':sales_report})




@login_required
def get_inventory_data(request):
    get_data = ProductInventory.objects.all()
    sale_cat = [label.product_category for label in get_data]
    
    get_cat = ProductInventory.objects.values('product_category').annotate(total_inventory =Sum('qty_sqft'))
    labels = [items['product_category'] for items in get_cat]
    inventory = [items['total_inventory'] for items in get_cat]
    # print(get_cat)
    # print(sale_cat)
    # print(labels)
    # print(inventory)
    
    current_date = date.today()-timedelta(11)
    # get_sales_report = customers_order_from_website.objects.filter(created_at__gt=current_date)
    get_sales_report = customers_order_from_website.objects.all()
    last_three_day_sales_report = get_sales_report.values('created_at__date').annotate(total_orders=Count('order_number')).order_by('created_at__date')
    print('************************************************')
    print(last_three_day_sales_report)
    print('************************************************')
    sales_list_label = [items['created_at__date'] for items in last_three_day_sales_report] 
    sales_list_count = [items['total_orders'] for items in last_three_day_sales_report] 
    
    return JsonResponse(data={
        'categories': labels,
        'inventories': inventory,
        'sales_date':sales_list_label,
        'sales_order_count':sales_list_count
    })
    



# def cart(request,product):
#     cart_list =[]
#     try:
#         getproduct = ProductInventory.objects.get(product_code=product)
#         cart_list.append(getproduct)
#         print(cart_list)
#         return render(request,'cart.html',{'cart_product':cart_list})
#     except getproduct.DoesNotExist:
#         return HttpResponse('Product not found')


    
    
def add_to_cart(request, product):
            
    if 'cart_list' not in request.session:
        request.session['cart_list'] = []
        

    try:
        getproduct = ProductInventory.objects.get(product_code=product)
        product_qty = request.POST.get(f'quantity_{product}')
        # print(f'66666666666666 {product_qty} &&&&&&&&&&&&&&&&&&&')
        
        cart_list = request.session['cart_list']
        print(cart_list)
        
        # Check if product is already in cart_list
        if getproduct.product_code not in [item['product_code'] for item in cart_list]:
            cart_list.append({
                'product_code': getproduct.product_code,
                'product_name': getproduct.product_name,
                'product_description': getproduct.product_description,
                'selling_price_sqft': getproduct.selling_price_sqft,
                'product_image': str(getproduct.product_image),
                'product_qty':product_qty,
                'available_qty_godown':getproduct.qty_sqft
                
            })
        
        request.session['cart_list'] = cart_list
        print(f'========= {cart_list}==============')
        messages.success(request, f"{getproduct.product_name} added to cart successfully!")
        return redirect('homepage')
    except ProductInventory.DoesNotExist:
        return HttpResponse('Product not found')
    
    
    

@require_POST
def remove_from_cart(request, product_code):
    if 'cart_list' in request.session:
        cart_list = request.session['cart_list']
        new_cart_list = [item for item in cart_list if item['product_code'] != product_code]
        
        if len(cart_list) == len(new_cart_list):
            return JsonResponse({'status': 'error', 'message': 'Product not found in cart'}, status=404)
        
        request.session['cart_list'] = new_cart_list
        print(cart_list)
        print('=============')
        print(request.session['cart_list'])
        print('===============')
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Cart not found'}, status=404)

    
    
def cart(request):

    
    if 'cart_list' not in request.session:
        request.session['cart_list']=[]
    
        # return HttpResponse('Error')
    
    # elif len(request.session['cart_list'])==0:
    #     return HttpResponse('Empty Cart')
    
    # client = razorpay.Client(auth=(settings.KEY,settings.SECRET))
    # print(f'=========== {client}===============')
    # payment = client.order.create(data={'amount':'500' , 'currency':'INR','payment_capture':1})
    # print('++++++++++++++')
    # print(payment)
    # print('++++++++++++++')
    # client = razorpay.Client(auth=(settings.KEY,settings.SECRET))
    # print(f'=========== {client}===============')
    # payment = client.order.create(data={'amount':'500' , 'currency':'INR','payment_capture':1})
    # print('++++++++++++++')
    # print(payment)
    # print('++++++++++++++')  
    
    # data = {'payment':payment,'cart_products': cart_list}
    cart_list = request.session['cart_list']    
    username = request.user
    print('<<<<<<<<<<<<<<<<<<<')
    print(username)
    print('<<<<<<<<<<<<<<<<<<<')
    get_user_details = CustomUser.objects.get(email=username)
    print(get_user_details.profile_img)
    
    data = {'user_profile':get_user_details,'cart_products': cart_list}
    
    
    
    
    if request.method == 'POST':
        print('========== inside POST method of purchase cart ===================')
        # return HttpResponse(request.POST)
        cart_list = request.session['cart_list']
        print(cart_list)
        
        customer_name = request.POST.get('customer_name')
        customer_email = request.POST.get('customer_email')
        customer_phone = request.POST.get('customer_phone')
        address = request.POST.get('shipping_address')
        pin_code = request.POST.get('pincode')
      
        
        # Extracting the product_code values
        product_codes = [item['product_code'] for item in cart_list]
        # Save purchase history or perform other actions
        order_items = []
        if len(product_codes)==0:
            return HttpResponse('No products selected')
        
        
        total_amount=0
        for product_code in product_codes:
            get_product_inventory = ProductInventory.objects.get(product_code=product_code)
            get_product_inventory.qty_sqft = get_product_inventory.qty_sqft -  float(request.POST.get(f'quantity_{product_code}'))
            
            get_product_inventory.save()
            total = float(request.POST.get(f'quantity_{product_code}'))*float(request.POST.get(f'selling_price_sqft_{product_code}'))
            total_amount+=total
            
            
        
        
        key = getattr(settings, 'KEY', None)
        secret = getattr(settings, 'SECRET', None)

        print(f'KEY: {key}')
        print(f'SECRET: {secret}')
        
        total_payment_amount = int(math.ceil(total_amount*100))
        
        client = razorpay.Client(auth=(settings.KEY,settings.SECRET))
        print(f'=========== {client}===============')
        payment = client.order.create(data={'amount':total_payment_amount , 'currency':'INR','payment_capture':1})
        print('++++++++++++++')
        print(payment)
        print('++++++++++++++')  
        
        data = {'payment':payment , 'total_payment_amount':total_payment_amount,'payment_id':payment['id']}
        response = {'order_id':request.POST.get('razorpay_order_id') , 'payment_id':request.POST.get('razorpay_payment_id'),'payment_signature':request.POST.get('razorpay_signature')}
        print('========== Payment Done succussfully ============== ')
        print(response)
        print('====================================================')
            
            
        # Create and save order
        order = customers_order_from_website(
            user_name = username,
            customer_name=customer_name,
            customer_email=customer_email,
            customer_phone=customer_phone,
            address=address,
            pin_code=pin_code,
            total_amount=total_amount,
            order_id=payment['id'],
            payment_id=response['payment_id'],
            is_paid = True
        )
        order.save()
        
        
        for product_code in product_codes:
            
            gross_amount = float(request.POST.get(f'quantity_{product_code}'))*float(request.POST.get(f'selling_price_sqft_{product_code}'))
            order_items.append({
                'order_number':order.order_number,
                'product_name': request.POST.get(f'product_name_{product_code}'),
                'product_code': request.POST.get(f'product_code_{product_code}'),
                'selling_price': request.POST.get(f'selling_price_sqft_{product_code}'),
                'product_qty' :request.POST.get(f'quantity_{product_code}'),
                'total_amount':total_amount,
                'gross_amount':gross_amount,
                'pin_code':pin_code
                })
            
   
            orderItem = website_orderItems(
                
                order_number_id = order.order_number,
                product_code_id = request.POST.get(f'product_code_{product_code}'),
                selling_price = request.POST.get(f'selling_price_sqft_{product_code}'),
                qty = request.POST.get(f'quantity_{product_code}'),
                gross_amount=gross_amount
                )
            orderItem.save()
        
        print(f'############### {order_items} ################# ')
        send_order_confirmation_email(order, order_items)
        request.session['cart_list']=[]
        
        success_payment_page = render_to_string('success_payment.html',{
                'order_id': order.order_id,
                'payment_id': order.payment_id,
                'user_profile':get_user_details
            
        })
        
        return HttpResponse(success_payment_page)
        # return render(request,'cart.html',data)
        # return HttpResponse(f'Purchase successful! {order_items}')  # Replace with redirect or success message as needed
    
    print('========== outside POST method of purchase cart ===================')
    return render(request, 'cart.html', data)
        
    
    
def send_order_confirmation_email(order, order_items):
    subject = 'Order Confirmation'
    order_email_template = render_to_string('orderEmailTemplate.html', {
        'customer_name': order.customer_name,
        'order_number': order.order_number,
        'order_items': order_items,
        'total_amount': order.total_amount,
        'address':order.address,
        'pin_code': order.pin_code,
    })
    plain_message = strip_tags(order_email_template)
    to = order.customer_email
    subject='Order Confirmation'
    send_mail(
                    subject,
                    '',
                    from_email=settings.DEFAULT_FROM_EMAIL,  # Sender's email (set in your settings)
                    recipient_list=[to],
                    html_message=order_email_template,
                    fail_silently=False,  # Set to True to suppress exceptions if sending fails
                )
    return HttpResponse(order_email_template)





def purchase_cart(request):
    
    
    
    if request.method == 'POST':
        print('========== inside POST method of purchase cart ===================')
        # return HttpResponse(request.POST)
        cart_list = request.session['cart_list']
        customer_name = request.POST.get('customer_name')
        customer_email = request.POST.get('customer_email')
        customer_phone = request.POST.get('customer_phone')
        address = request.POST.get('shipping_address')
        pin_code = request.POST.get('pincode')
      
        
        # Extracting the product_code values
        product_codes = [item['product_code'] for item in cart_list]
        # Save purchase history or perform other actions
        order_items = []
        if len(product_codes)==0:
            return HttpResponse('No products selected')
        
        
        total_amount=0
        for product_code in product_codes:
            get_product_inventory = ProductInventory.objects.get(product_code=product_code)
            get_product_inventory.qty_sqft = get_product_inventory.qty_sqft -  float(request.POST.get(f'quantity_{product_code}'))
            
            get_product_inventory.save()
            total = float(request.POST.get(f'quantity_{product_code}'))*float(request.POST.get(f'selling_price_sqft_{product_code}'))
            total_amount+=total
            
            
            
        # Create and save order
        order = customers_order_from_website(
            customer_name=customer_name,
            customer_email=customer_email,
            customer_phone=customer_phone,
            address=address,
            pin_code=pin_code,
            total_amount=total_amount
        )
        order.save()
        
        
        for product_code in product_codes:
            
            gross_amount = float(request.POST.get(f'quantity_{product_code}'))*float(request.POST.get(f'selling_price_sqft_{product_code}'))
            order_items.append({
                'order_number':order.order_number,
                'product_name': request.POST.get(f'product_name_{product_code}'),
                'product_code': request.POST.get(f'product_code_{product_code}'),
                'selling_price': request.POST.get(f'selling_price_sqft_{product_code}'),
                'product_qty' :request.POST.get(f'quantity_{product_code}'),
                'total_amount':total_amount,
                'gross_amount':gross_amount,
                'pin_code':pin_code
                })
            
   
            orderItem = website_orderItems(
                
                order_number_id = order.order_number,
                product_code_id = request.POST.get(f'product_code_{product_code}'),
                selling_price = request.POST.get(f'selling_price_sqft_{product_code}'),
                qty = request.POST.get(f'quantity_{product_code}'),
                gross_amount=gross_amount
                )
            orderItem.save()
            
        key = getattr(settings, 'KEY', None)
        secret = getattr(settings, 'SECRET', None)

        print(f'KEY: {key}')
        print(f'SECRET: {secret}')
        
        total_payment_amount = order.total_amount * 100
        
        client = razorpay.Client(auth=(settings.KEY,settings.SECRET))
        print(f'=========== {client}===============')
        payment = client.order.create(data={'amount':total_payment_amount , 'currency':'INR','payment_capture':1})
        print('++++++++++++++')
        print(payment)
        print('++++++++++++++')  
        
        data = {'payment':payment , 'total_payment_amount':total_payment_amount}
        # send_order_confirmation_email(order, order_items)
        request.session['cart_list']=[]

        return render(request,'order_placed.html',data)
        # return HttpResponse(f'Purchase successful! {order_items}')  # Replace with redirect or success message as needed
    
    return HttpResponse('Method not allowed', status=405)





def order_placed(request):
    message_list = messages.success(request, f"Yeaahh !!! Your Order Placed successfully!")

    return render(request,'order_placed.html',{'message_list':message_list})


def notify_website_order_to_admin(request):
    
    get_unaccepted_orders = customers_order_from_website.objects.filter(is_accepted=False)
    get_accepted_orders = customers_order_from_website.objects.filter(is_accepted=True)
    
    # Example of fetching related order items for each order
    order_items_dict = {}   
    for order in get_unaccepted_orders:
        items = website_orderItems.objects.filter(order_number=order)
        order_items_dict[order] = items
    
    print('***********************************')
    print(get_unaccepted_orders)
    print(order_items_dict)
    
    print('***********************************')
    if 'accept_order_id' in request.GET:
        order_number = request.GET.get('accept_order_id')
        print(f'============ {order_number} =============== ')
        accept_order = customers_order_from_website.objects.get(order_number=order_number)
        accept_order.is_accepted=True
        accept_order.save()
        return redirect('websiteorders')
        
        
        
    # if request.GET.get('download')=='xlsx':

    #     wb = Workbook()
    #     ws = wb.active

    #     # Add headers to the worksheet
    #     ws.append(['Order Number','Product Code','Customer Name', 'Customer Address', 'Mobile Number' ,'Total Sqft', 'Per sqft price' ,'Total Bill Amount'])

    #     # Add data rows to the worksheet
    #     for order in get_accepted_orders:
    #         ws.append([order.order_number,order.created_at,order.customer_name, order.address, order.customer_phone, order.customer_email, order.total_amount, order.order_id,order.payment_id])

    #     # Save the workbook to a BytesIO object
    #     output = io.BytesIO()
    #     wb.save(output)
    #     output.seek(0)

    #     # Create an HTTP response with the Excel file as attachment
    #     response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #     response['Content-Disposition'] = f'attachment; filename="sales_report {current_date}.xlsx"'
    #     print('sales file created ')
    #     return response
    
    
    return render(request,'notify_orders.html',{'orders':get_unaccepted_orders,'order_items': order_items_dict})



@csrf_exempt
def queryform_data(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        mobile = request.POST.get('phone')
        msg = request.POST.get('message')
        
        store_query = Queryform.objects.create(name=name, mobile_number=mobile, message=msg)
        store_query.save()
        
        return redirect('homepage') 
    
        
        
        
def user_past_purchase(request):
    
    username = request.user
    get_user_details = CustomUser.objects.get(email=username)    
    
    get_user_purchase = customers_order_from_website.objects.get(user_name=username)
    get_user_purchase_item = website_orderItems.objects.filter(order_number_id=get_user_purchase.order_number)
    print(get_user_purchase_item)
    
    past_purchase = []
    
    for item in get_user_purchase_item:
        product = ProductInventory.objects.get(product_code=item.product_code_id)
        past_purchase.append({
            'order_number':get_user_purchase.order_id,
            'product_name':product.product_name,
            'qty':item.qty,
            'price':get_user_purchase.total_amount,
            'date':get_user_purchase.created_at,
            'product_img':product.product_image
            
        })
        
    print(past_purchase)
    return render(request,'past_user_purchase.html',{'past_purchase':past_purchase,'user_profile':get_user_details})
    